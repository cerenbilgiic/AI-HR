import csv
import io

from openpyxl import Workbook, load_workbook
from pydantic import EmailStr, TypeAdapter, ValidationError
from sqlalchemy.orm import Session

from app.models.audit_log import AuditLog
from app.models.candidate import Candidate
from app.models.job import Job
from app.schemas.candidate import CandidateCreate
from app.schemas.invitation import CandidateImportSummary, ImportRowDuplicate, ImportRowError
from app.services import candidate_service

REQUIRED_COLUMNS = ["Ad", "Soyad", "E-posta", "Pozisyon"]
TEMPLATE_COLUMNS = ["Ad", "Soyad", "E-posta", "Telefon", "Pozisyon", "Departman"]

_email_adapter = TypeAdapter(EmailStr)


def build_template_xlsx() -> io.BytesIO:
    """.xlsx, not .csv — a comma-delimited CSV opens as a single column in
    Excel on any locale where the list separator isn't a comma (e.g.
    Turkish Windows uses ';'). A real workbook has no delimiter to get
    wrong, so columns always render correctly regardless of locale."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(TEMPLATE_COLUMNS)
    for column_cells in sheet.columns:
        header_length = len(str(column_cells[0].value))
        sheet.column_dimensions[column_cells[0].column_letter].width = max(header_length + 4, 14)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _rows_from_csv(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter, [])]
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        rows.append({header[i]: ("" if v is None else str(v)) for i, v in enumerate(values) if i < len(header)})
    return rows


def parse_and_import(
    db: Session, filename: str, data: bytes, actor_id: int | None = None
) -> CandidateImportSummary:
    """Row-by-row CSV/XLSX import — see /candidates/import. One malformed
    row is reported and skipped, never aborts the whole file. Duplicate
    check covers both the existing DB (Candidate.email is unique) and
    repeats within the same file."""
    if filename.lower().endswith(".xlsx"):
        rows = _rows_from_xlsx(data)
    else:
        rows = _rows_from_csv(data)

    existing_emails = {e.lower() for (e,) in db.query(Candidate.email).all()}
    jobs_by_title = {j.title.strip().lower(): j for j in db.query(Job).all()}

    created = 0
    errors: list[ImportRowError] = []
    duplicates: list[ImportRowDuplicate] = []
    seen_in_file: set[str] = set()

    for i, row in enumerate(rows):
        row_number = i + 2  # header is row 1, data starts at row 2
        first_name = (row.get("Ad") or "").strip()
        last_name = (row.get("Soyad") or "").strip()
        email = (row.get("E-posta") or "").strip()
        phone = (row.get("Telefon") or "").strip() or None
        position = (row.get("Pozisyon") or "").strip()

        missing = [
            label
            for label, value in [("Ad", first_name), ("Soyad", last_name), ("E-posta", email), ("Pozisyon", position)]
            if not value
        ]
        if missing:
            errors.append(ImportRowError(row=row_number, message=f"Eksik alan(lar): {', '.join(missing)}"))
            continue

        try:
            _email_adapter.validate_python(email)
        except ValidationError:
            errors.append(ImportRowError(row=row_number, message=f"Geçersiz e-posta formatı: {email}"))
            continue

        email_key = email.lower()
        if email_key in existing_emails or email_key in seen_in_file:
            duplicates.append(ImportRowDuplicate(row=row_number, email=email))
            continue

        job = jobs_by_title.get(position.lower())
        if job is None:
            errors.append(ImportRowError(row=row_number, message=f"Pozisyon bulunamadı: {position}"))
            continue

        candidate_service.create_candidate(
            db,
            CandidateCreate(
                full_name=f"{first_name} {last_name}", email=email, phone=phone, job_id=job.id, password=None
            ),
        )
        seen_in_file.add(email_key)
        created += 1

    db.add(
        AuditLog(
            actor_type="hr",
            actor_id=actor_id,
            action="candidates_imported",
            detail={"created": created, "errors": len(errors), "duplicates": len(duplicates)},
        )
    )
    db.commit()

    return CandidateImportSummary(created=created, errors=errors, duplicates=duplicates)
