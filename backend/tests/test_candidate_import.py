import csv
import io

from openpyxl import Workbook, load_workbook

from app.core.security import create_access_token
from app.models.candidate import Candidate
from app.services import candidate_import_service


def _csv_bytes(rows: list[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=candidate_import_service.TEMPLATE_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def test_import_template_has_expected_columns():
    buffer = candidate_import_service.build_template_xlsx()
    sheet = load_workbook(buffer).active
    header = [cell.value for cell in next(sheet.iter_rows())]
    assert header == ["Ad", "Soyad", "E-posta", "Telefon", "Pozisyon", "Departman"]


def test_download_import_template(client, as_hr):
    resp = client.get("/api/v1/candidates/import/template")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_import_csv_creates_valid_candidates(db_session, job):
    data = _csv_bytes(
        [{"Ad": "Elif", "Soyad": "Kara", "E-posta": "elif.kara.import@example.com", "Pozisyon": job.title}]
    )

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 1
    assert summary.errors == []
    assert summary.duplicates == []
    created = db_session.query(Candidate).filter(Candidate.email == "elif.kara.import@example.com").first()
    assert created is not None
    assert created.job_id == job.id
    assert created.hashed_password is None
    assert created.phone is None


def test_import_csv_stores_phone_when_present(db_session, job):
    data = _csv_bytes(
        [
            {
                "Ad": "Elif",
                "Soyad": "Kara",
                "E-posta": "elif.kara.import@example.com",
                "Telefon": "+90 555 111 2233",
                "Pozisyon": job.title,
            }
        ]
    )

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 1
    created = db_session.query(Candidate).filter(Candidate.email == "elif.kara.import@example.com").first()
    assert created.phone == "+90 555 111 2233"


def test_import_csv_reports_missing_fields(db_session, job):
    data = _csv_bytes([{"Soyad": "Kara", "E-posta": "elif.kara.import@example.com", "Pozisyon": job.title}])

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 0
    assert len(summary.errors) == 1
    assert summary.errors[0].row == 2
    assert "Ad" in summary.errors[0].message


def test_import_csv_reports_bad_email(db_session, job):
    data = _csv_bytes([{"Ad": "Elif", "Soyad": "Kara", "E-posta": "not-an-email", "Pozisyon": job.title}])

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 0
    assert len(summary.errors) == 1
    assert "e-posta" in summary.errors[0].message.lower()


def test_import_csv_reports_unknown_position(db_session, job):
    data = _csv_bytes(
        [
            {
                "Ad": "Elif",
                "Soyad": "Kara",
                "E-posta": "elif.kara.import@example.com",
                "Pozisyon": "Nonexistent Job Title Xyz",
            }
        ]
    )

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 0
    assert len(summary.errors) == 1
    assert "Pozisyon" in summary.errors[0].message


def test_import_csv_reports_duplicate_existing_email(db_session, job, candidate):
    data = _csv_bytes([{"Ad": "Elif", "Soyad": "Kara", "E-posta": candidate.email, "Pozisyon": job.title}])

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 0
    assert len(summary.duplicates) == 1
    assert summary.duplicates[0].email == candidate.email


def test_import_csv_reports_duplicate_within_file(db_session, job):
    data = _csv_bytes(
        [
            {"Ad": "Elif", "Soyad": "Kara", "E-posta": "elif.kara.import@example.com", "Pozisyon": job.title},
            {"Ad": "Elif", "Soyad": "Kara2", "E-posta": "elif.kara.import@example.com", "Pozisyon": job.title},
        ]
    )

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.csv", data)

    assert summary.created == 1
    assert len(summary.duplicates) == 1


def test_import_endpoint_requires_hr_auth(client, candidate, job):
    # Real JWT decoding path, not the as_candidate fixture — its blanket
    # get_current_user override bypasses the dependency's own type check
    # (see test_hr_dashboard.py::test_hr_endpoints_reject_candidate_token).
    token = create_access_token(subject=str(candidate.id), token_type="candidate")
    data = _csv_bytes(
        [{"Ad": "Elif", "Soyad": "Kara", "E-posta": "elif.kara.import@example.com", "Pozisyon": job.title}]
    )
    resp = client.post(
        "/api/v1/candidates/import",
        files={"file": ("adaylar.csv", data, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 401


def test_import_endpoint_end_to_end(client, as_hr, job):
    data = _csv_bytes(
        [{"Ad": "Elif", "Soyad": "Kara", "E-posta": "elif.kara.import@example.com", "Pozisyon": job.title}]
    )

    resp = client.post("/api/v1/candidates/import", files={"file": ("adaylar.csv", data, "text/csv")})

    assert resp.status_code == 200
    body = resp.json()
    assert body["created"] == 1
    assert body["errors"] == []
    assert body["duplicates"] == []


def test_import_xlsx_creates_valid_candidates(db_session, job):
    wb = Workbook()
    ws = wb.active
    ws.append(["Ad", "Soyad", "E-posta", "Telefon", "Pozisyon", "Departman"])
    ws.append(["Mert", "Aydın", "mert.aydin.import@example.com", "+90 555 999 1122", job.title, job.department or ""])
    buffer = io.BytesIO()
    wb.save(buffer)

    summary = candidate_import_service.parse_and_import(db_session, "adaylar.xlsx", buffer.getvalue())

    assert summary.created == 1
    created = db_session.query(Candidate).filter(Candidate.email == "mert.aydin.import@example.com").first()
    assert created is not None
    assert created.phone == "+90 555 999 1122"
